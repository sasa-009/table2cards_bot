from openpyxl import load_workbook

from data import update_data


def convert():
    wb = load_workbook(u''+'./downloads/data.xlsx')
    sheet_names = wb.sheetnames
    
    data = {"config":{
    "transc": False,
    "quantity_words": 3
}}
    word_id = 0


    for i in sheet_names:
        sheet = wb[i]
        for cellObj in sheet['A1':'C' + str(sheet.max_row)]:
            word = {
                "word": None,
                "transc": None,
                "transl": None,
                "learn": None,
                "rd": None,
                "rdl": None,
                "known": None,
                "tags": [],
            }
            for cells in cellObj:
                value = cells.value
                if type(value) == str:
                    value = value.strip()
                if cells.column == 1:
                    word["word"] = value
                elif cells.column == 2:
                    if cells.value == None:
                        word = {}
                        continue
                    else:
                        word["transc"] = value
                elif cells.column == 3:
                    if cells.value == None:
                        word = {}
                        continue
                    else:
                        word["transl"] = value
            if word != {}:
                word_id +=1
                data[word_id] = word
    update_data(data)

if __name__ == "__main__":
    convert()